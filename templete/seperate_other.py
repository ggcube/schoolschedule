import openpyxl, os, sys, datetime, csv
temp_ver ='1.0' # 樣板版本
def text2num(t):
    if t=='一':
        return 1
    if t=='二':
        return 2
    if t=='三':
        return 3
    if t=='四':
        return 4
    if t=='五':
        return 5
    if t=='六':
        return 6
    if t=='七':
        return 7
    if t=='八':
        return 8
    if t=='九':
        return 9
    return 0



# # 讀取原課表檔案
# wb = openpyxl.open(os.path.join('input',os.listdir('input')[0])) 
# sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
# 讀取樣板文件
templete_dir = os.path.join('templete','class_schedule_v'+temp_ver+'.xlsx')
new_wb = openpyxl.open(templete_dir)
class_sheet = new_wb.create_sheet()
if 'class_sheet' in new_wb.get_sheet_names():
    new_wb.remove_sheet(new_wb.get_sheet_by_name('class_sheet')) 
class_sheet.title = 'class_sheet' # 工作表改名成預設
# 生成輸出路徑
if not(os.path.isdir('output')):
    os.makedirs('output')
output_dir = os.path.join('output',datetime.datetime.now().strftime('%Y%m%d-%H%M%S')+'.xlsx')

# # 開目標 xlsx 物件(或新增)
# if len(os.listdir('output')) == 0:
#     print('\n目標路徑無試算表檔案，啟用新增模式。\n')
#     output_dir = os.path.join('output','class_schedule.xlsx')
#     new_wb = openpyxl.Workbook() # 建立新的活頁簿
#     class_sheet = new_wb.get_sheet_by_name(new_wb.get_sheet_names()[0]) 
#     class_sheet.title = 'class_sheet' # 工作表改名成預設
# elif len(os.listdir('output')) == 1:
#     print('\n找到目標試算表檔案，啟用覆寫模式。\n')
#     output_dir = os.path.join('output',os.listdir('output')[0])
#     new_wb = openpyxl.open(output_dir)
#     class_sheet = new_wb.create_sheet()
#     if 'class_sheet' in new_wb.get_sheet_names():
#         new_wb.remove_sheet(new_wb.get_sheet_by_name('class_sheet')) 
#     class_sheet.title = 'class_sheet' # 工作表改名成預設
# else:
#     print('\n找到2個以上檔案，請移除。\n')
#     sys.exit()


# 製作以班級為主鍵的課務資料
    # 寫入標題
class_title = ['班級','星期','堂次','課程名稱','教師']
for i in range(5):
    class_sheet.cell(row=1,column=i+1).value = class_title[i]
data_num = 1 # 第一列寫入標題了，從第二列開始。
# 讀取原課表檔案 '.csv' 並寫入樣版檔
with open(os.path.join('input',os.listdir('input')[0]), newline='') as csvfile:
    rows = csv.reader(csvfile)
    for row in rows:
        if data_num == 1:
            data_num += 1
            continue
        class_sheet.cell(row=data_num,column=1).value = text2num(row[2][0])*100+int(row[3][1:3])
        class_sheet.cell(row=data_num,column=2).value = text2num(row[0][1]) 
        class_sheet.cell(row=data_num,column=3).value = text2num(row[1][1]) 
        class_sheet.cell(row=data_num,column=4).value = row[5]
        class_sheet.cell(row=data_num,column=5).value = row[4]
        data_num += 1
         
# 存檔結束
new_wb.save(output_dir)
print('已產生課表excel檔案...')
os.system('pause')