import openpyxl, os, sys, datetime
temp_ver ='1.0' # 樣板版本
# 讀取原課表檔案
wb = openpyxl.open(os.path.join('input',os.listdir('input')[0])) 
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
# 讀取樣板文件
templete_dir = os.path.join('templete','class_schedule_v'+temp_ver+'.xlsx')
new_wb = openpyxl.open(templete_dir)
class_sheet = new_wb.create_sheet()
if 'class_sheet' in new_wb.get_sheet_names():
    new_wb.remove_sheet(new_wb.get_sheet_by_name('class_sheet')) 
class_sheet.title = 'class_sheet' # 工作表改名成預設
# 生成輸出路徑
if not(os.path.isdir('output')):
    os.makedirs('output')
output_dir = os.path.join('output',datetime.datetime.now().strftime('%Y%m%d-%H%M%S')+'.xlsx')


# # 開目標 xlsx 物件(或新增)
# if len(os.listdir('output')) == 0:
#     print('\n目標路徑無試算表檔案，啟用新增模式。\n')
#     output_dir = os.path.join('output','class_schedule.xlsx')
#     new_wb = openpyxl.Workbook() # 建立新的活頁簿
#     class_sheet = new_wb.get_sheet_by_name(new_wb.get_sheet_names()[0]) 
#     class_sheet.title = 'class_sheet' # 工作表改名成預設
# elif len(os.listdir('output')) == 1:
#     print('\n找到目標試算表檔案，啟用覆寫模式。\n')
#     output_dir = os.path.join('output',os.listdir('output')[0])
#     new_wb = openpyxl.open(output_dir)
#     class_sheet = new_wb.create_sheet()
#     if 'class_sheet' in new_wb.get_sheet_names():
#         new_wb.remove_sheet(new_wb.get_sheet_by_name('class_sheet')) 
#     class_sheet.title = 'class_sheet' # 工作表改名成預設
# else:
#     print('\n找到2個以上檔案，請移除。\n')
#     sys.exit()


# 解析課表檔案
class_count = sheet.max_row - 1 # 班級數
'''
    # 星期 int((column-3)/7+1)
    # 堂數 (column-3)%7+1 
    # 解析單一儲存格
        # cell_data = sheet['C2'].value.split('_')
        # 第一格是課程名稱，第三格是教師姓名
'''
# 製作以班級為主鍵的課務資料
    # 寫入標題
class_title = ['班級','星期','堂次','課程名稱','教師']
for i in range(5):
    class_sheet.cell(row=1,column=i+1).value = class_title[i]
    # 寫入資料
data_num = 2 # 第一列寫入標題了，從第二列開始。
for r in range(2,sheet.max_row + 1):
    for c in range(3,sheet.max_column+1):
        cell_data = sheet.cell(row=r,column=c).value.split('_')
        if len(cell_data) >= 2 : # 空的儲存格不會執行
            class_sheet.cell(row=data_num,column=1).value = int(sheet.cell(row=r,column=1).value)
            class_sheet.cell(row=data_num,column=2).value = int(sheet.cell(row=1,column=c).value[1])
            class_sheet.cell(row=data_num,column=3).value = int(sheet.cell(row=1,column=c).value[2]) 
            class_sheet.cell(row=data_num,column=4).value = cell_data[0]
            class_sheet.cell(row=data_num,column=5).value = cell_data[2]
            # print('test', sheet.cell(row=r,column=5).value)
            data_num += 1
            
# 存檔結束
new_wb.save(output_dir)
print('課表檔案已產製...')
os.system('pause')